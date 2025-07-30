import streamlit as st
import pandas as pd
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import os
from datetime import datetime
import io
import tempfile
import sys
# Add the current directory to the Python path
current_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(current_dir)

# Import the page modules (will work after you create the files)
try:
    from pages.about_us import show_about_page
    from pages.privacy_policy import show_privacy_policy
    from pages.contact_us import show_contact_page
    from utils.helpers import apply_custom_css as custom_css
except ImportError:
    # Fallback if files don't exist yet
    def show_about_page():
        st.write("About Us page - Create pages/about_us.py file")
    def show_privacy_policy():
        st.write("Privacy Policy page - Create pages/privacy_policy.py file")
    def show_contact_page():
        st.write("Contact Us page - Create pages/contact_us.py file")
    def custom_css():
        pass

# --- Enhanced Custom CSS for Modern UI ---
def apply_custom_css():
    st.markdown("""
    <style>
    /* Import modern font */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    /* Global styles */
    .stApp {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
        background: white !important;
        min-height: 100vh;
    }

    
    /* Main container with glass effect */
    .main .block-container {
        background: rgba(255, 255, 255, 0.95) !important;
        border-radius: 20px !important;
        box-shadow: 0 20px 40px rgba(0, 0, 0, 0.08) !important;
        padding: 2rem !important;
        margin: 1rem auto !important;
        transition: all 0.4s ease !important;
        animation: smoothFadeIn 0.7s ease-out forwards !important;
        transform: translateY(20px);
        opacity: 0;
    }

    @keyframes smoothFadeIn {
        to {
            transform: translateY(0);
            opacity: 1;
        }
    }

    
    /* Modern title with gradient */
    .main-title {
        font-family: 'Inter', sans-serif !important;
        font-size: 3.5rem !important;
        font-weight: 700 !important;
        text-align: center;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 50%, #f093fb 100%) !important;
        background-size: 200% 200% !important;
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
        animation: gradientFlow 4s ease-in-out infinite !important;
        margin-bottom: 1rem !important;
        letter-spacing: -2px !important;
    }
    
    @keyframes gradientFlow {
        0%, 100% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
    }
    
    /* Subtle subtitle */
    .subtitle {
        font-family: 'Inter', sans-serif !important;
        font-size: 1.2rem !important;
        font-weight: 400 !important;
        text-align: center;
        color: #64748b !important;
        margin-bottom: 2rem !important;
        animation: fadeInUp 1s ease-out 0.3s both !important;
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    /* Modern button styles - FIXED VERSION */
    .stButton > button {
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        border-radius: 12px !important;
        padding: 0.75rem 1.5rem !important;
        border: none !important;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3) !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        transform: translateY(0) !important;
    }

    .stButton > button:hover {
        transform: translateY(-2px) scale(1.02) !important;
        box-shadow: 0 8px 25px rgba(102, 126, 234, 0.4) !important;
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%) !important;
        color: white !important;
    }

    .stButton > button:active {
        transform: translateY(0) scale(0.98) !important;
        transition: transform 0.1s !important;
    }

    /* Primary button special styling - FIXED VERSION */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #06b6d4 0%, #3b82f6 100%) !important;
        color: white !important;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.3) !important;
    }

    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #3b82f6 0%, #06b6d4 100%) !important;
        color: white !important;
        box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4) !important;
        transform: translateY(-2px) scale(1.02) !important;
    
    /* Success messages with modern styling - FIXED VERSION */
    .success-message {
        background: linear-gradient(135deg, #10b981 0%, #059669 100%) !important;
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 1rem 1.5rem !important;
        border-radius: 12px !important;
        margin: 1rem 0 !important;
        box-shadow: 0 4px 15px rgba(16, 185, 129, 0.2) !important;
        animation: slideInLeft 0.5s ease-out !important;
        border-left: 4px solid #065f46 !important;
        transition: all 0.3s ease !important;
    }

    .success-message:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(16, 185, 129, 0.3) !important;
    }

    /* Info messages - FIXED VERSION */
    .info-message {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%) !important;
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 1rem 1.5rem !important;
        border-radius: 12px !important;
        margin: 1rem 0 !important;
        box-shadow: 0 4px 15px rgba(59, 130, 246, 0.2) !important;
        animation: slideInRight 0.5s ease-out !important;
        border-left: 4px solid #1e40af !important;
        transition: all 0.3s ease !important;
    }

    .info-message:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(59, 130, 246, 0.3) !important;
    }

    /* Error messages - FIXED VERSION */
    .error-message {
        background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%) !important;
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 1rem 1.5rem !important;
        border-radius: 12px !important;
        margin: 1rem 0 !important;
        box-shadow: 0 4px 15px rgba(239, 68, 68, 0.2) !important;
        animation: shake 0.5s ease-in-out !important;
        border-left: 4px solid #b91c1c !important;
        transition: all 0.3s ease !important;
    }

    .error-message:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(239, 68, 68, 0.3) !important;
    }

    /* Warning messages - FIXED VERSION */
    .warning-message {
        background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%) !important;
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 1rem 1.5rem !important;
        border-radius: 12px !important;
        margin: 1rem 0 !important;
        box-shadow: 0 4px 15px rgba(245, 158, 11, 0.2) !important;
        animation: slideInUp 0.5s ease-out !important;
        border-left: 4px solid #b45309 !important;
        transition: all 0.3s ease !important;
    }

    .warning-message:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 20px rgba(245, 158, 11, 0.3) !important;
    }
    
    /* Modern progress bar */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #06b6d4, #3b82f6, #8b5cf6) !important;
        border-radius: 8px !important;
        animation: progressGlow 2s ease-in-out infinite alternate !important;
    }
    
    @keyframes progressGlow {
        from { box-shadow: 0 0 5px rgba(59, 130, 246, 0.5); }
        to { box-shadow: 0 0 20px rgba(59, 130, 246, 0.8); }
    }
    
    /* Modern tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.5rem;
        background: rgba(248, 250, 252, 0.8) !important;
        padding: 0.5rem !important;
        border-radius: 12px !important;
        backdrop-filter: blur(10px) !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        padding: 0.75rem 1.5rem !important;
        border-radius: 8px !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border: none !important;
        background: transparent !important;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: rgba(255, 255, 255, 0.7) !important;
        transform: translateY(-1px) !important;
    }
    
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: white !important;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1) !important;
        color: #667eea !important;
    }
    
    /* Modern dataframe styling - FIXED VERSION */
    .stDataFrame {
        animation: fadeInScale 0.6s ease-out !important;
        border-radius: 12px !important;
        overflow: hidden !important;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.1) !important;
        font-family: 'Inter', sans-serif !important;
        transition: all 0.3s ease !important;
    }

    .stDataFrame:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 30px rgba(0, 0, 0, 0.15) !important;
    }

    /* Dataframe headers */
    .stDataFrame thead th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        font-family: 'Inter', sans-serif !important;
        font-weight: 600 !important;
        padding: 12px !important;
    }

    /* Dataframe cells */
    .stDataFrame tbody td {
        font-family: 'Inter', sans-serif !important;
        color: #1e293b !important;
        padding: 10px !important;
        transition: background-color 0.2s ease !important;
    }

    .stDataFrame tbody tr:hover td {
        background-color: rgba(102, 126, 234, 0.05) !important;
    }
    
    @keyframes fadeInScale {
        from {
            opacity: 0;
            transform: scale(0.95);
        }
        to {
            opacity: 1;
            transform: scale(1);
        }
    }
    
    /* Modern headers */
    h1, h2, h3 {
        font-family: 'Inter', sans-serif !important;
        font-weight: 600 !important;
        color: #1e293b !important;
        animation: fadeInDown 0.6s ease-out !important;
    }
    
    @keyframes fadeInDown {
        from {
            opacity: 0;
            transform: translateY(-10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    /* Form containers */
    .stForm {
        background: rgba(255, 255, 255, 0.8) !important;
        backdrop-filter: blur(10px) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1) !important;
        animation: slideInScale 0.6s ease-out !important;
    }
    
    @keyframes slideInScale {
        from {
            opacity: 0;
            transform: scale(0.9) translateY(20px);
        }
        to {
            opacity: 1;
            transform: scale(1) translateY(0);
        }
    }
    
    /* Input fields - FIXED VERSION */
    .stNumberInput input, .stSelectbox > div > div, .stTextInput input {
        border-radius: 8px !important;
        border: 2px solid #e2e8f0 !important;
        transition: all 0.3s ease !important;
        font-family: 'Inter', sans-serif !important;
        color: #1e293b !important;
        background: white !important;
    }

    .stNumberInput input:focus, .stSelectbox > div > div:focus, .stTextInput input:focus {
        border-color: #667eea !important;
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1) !important;
        outline: none !important;
        transform: translateY(-1px) !important;
    }

    /* Selectbox styling */
    .stSelectbox > div > div {
        color: #1e293b !important;
        background: white !important;
    }

    .stSelectbox > div > div > div {
        color: #1e293b !important;
    }
    
    /* File uploader */
    .stFileUploader {
        background: rgba(255, 255, 255, 0.9) !important;
        border-radius: 12px !important;
        border: 2px dashed #cbd5e1 !important;
        padding: 2rem !important;
        transition: all 0.3s ease !important;
        animation: fadeIn 0.8s ease-out !important;
    }
    
    .stFileUploader:hover {
        border-color: #667eea !important;
        background: rgba(102, 126, 234, 0.05) !important;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        background: rgba(255, 255, 255, 0.9) !important;
        border-radius: 8px !important;
        transition: all 0.3s ease !important;
    }
    
    .streamlit-expanderHeader:hover {
        background: rgba(102, 126, 234, 0.1) !important;
        transform: translateX(5px) !important;
    }
    
    /* Spinner enhancement */
    .stSpinner {
        animation: modernSpin 1s linear infinite !important;
    }
    
    @keyframes modernSpin {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }
    
    /* Status cards - FIXED VERSION */
    .status-card {
        background: rgba(255, 255, 255, 0.95) !important;
        backdrop-filter: blur(10px) !important;
        border-radius: 12px !important;
        padding: 1.5rem !important;
        margin: 0.5rem 0 !important;
        border-left: 4px solid #10b981 !important;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1) !important;
        transition: all 0.3s ease !important;
        animation: slideInUp 0.5s ease-out !important;
        color: #1e293b !important;
        font-family: 'Inter', sans-serif !important;
    }

    .status-card:hover {
        transform: translateY(-2px) scale(1.01) !important;
        box-shadow: 0 8px 25px rgba(0, 0, 0, 0.15) !important;
    }

    .status-card strong {
        color: #1e293b !important;
        font-family: 'Inter', sans-serif !important;
    }

    
    /* Celebration animation */
    .celebration {
        animation: celebrate 1.5s ease-in-out !important;
    }
    
    @keyframes celebrate {
        0%, 100% { transform: scale(1); }
        25% { transform: scale(1.05) rotate(1deg); }
        50% { transform: scale(1.1) rotate(-1deg); }
        75% { transform: scale(1.05) rotate(1deg); }
    }
    
    /* Results container */
    .results-container {
        background: rgba(255, 255, 255, 0.95) !important;
        backdrop-filter: blur(15px) !important;
        border-radius: 16px !important;
        padding: 2rem !important;
        margin: 1rem 0 !important;
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1) !important;
        border: 1px solid rgba(255, 255, 255, 0.2) !important;
        animation: slideInUp 0.8s ease-out !important;
    }
    
    /* Column headers */
    .column-header {
        font-weight: 600 !important;
        color: #1e293b !important;
        font-size: 1.1rem !important;
        margin-bottom: 1rem !important;
    }
    
    /* Responsive adjustments */
    @media (max-width: 768px) {
        .main-title {
            font-size: 2.5rem !important;
        }
        
        .main .block-container {
            margin: 0.5rem !important;
            padding: 1rem !important;
        }
    }
    /* Smooth scroll and transition for all */
    html, body {
        scroll-behavior: smooth;
        transition: background-color 0.3s ease, color 0.3s ease;
    }

    * {
        transition: all 0.25s ease-in-out;
    }

    /* Smooth pop-in for result containers and cards */
    .results-container, .status-card {
        animation: popIn 0.5s ease-out forwards;
        transform: scale(0.95);
        opacity: 0;
    }

    @keyframes popIn {
        to {
            transform: scale(1);
            opacity: 1;
        }
    }

    /* Pop-up buttons */
    .stButton > button:hover {
        transform: scale(1.03) translateY(-2px) !important;
    }
    /* Enhanced smooth transitions for all interactive elements */
    .stButton, .stSelectbox, .stNumberInput, .stTextInput, .stFileUploader, 
    .stDataFrame, .success-message, .info-message, .error-message, .warning-message,
    .status-card, .results-container {
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
    }

    /* Smooth popup animations */
    @keyframes smoothPopIn {
        0% {
            opacity: 0;
            transform: scale(0.9) translateY(20px);
        }
        100% {
            opacity: 1;
            transform: scale(1) translateY(0);
        }
    }

    /* Apply popup animation to key elements */
    .stButton, .success-message, .info-message, .error-message, .warning-message {
        animation: smoothPopIn 0.4s ease-out !important;
    }

    /* Enhanced text contrast */
    h1, h2, h3, h4, h5, h6, p, span, div {
        color: #1e293b !important;
        font-family: 'Inter', sans-serif !important;
    }

    /* Ensure proper text visibility in all containers */
    .main .block-container, .results-container, .status-card {
        color: #1e293b !important;
    }

    /* Fix any remaining text visibility issues */
    .stMarkdown, .stText {
        color: #1e293b !important;
    }

    /* Column headers enhancement */
    .column-header {
        font-weight: 600 !important;
        color: #1e293b !important;
        font-size: 1.2rem !important;
        margin-bottom: 1rem !important;
        font-family: 'Inter', sans-serif !important;
        text-shadow: none !important;
    }                   
    </style>
    """, unsafe_allow_html=True)

# --- Enhanced Message Functions ---
def show_success_message(message):
    st.markdown(f'<div class="success-message">✅ {message}</div>', unsafe_allow_html=True)

def show_info_message(message):
    st.markdown(f'<div class="info-message">ℹ️ {message}</div>', unsafe_allow_html=True)

def show_error_message(message):
    st.markdown(f'<div class="error-message">❌ {message}</div>', unsafe_allow_html=True)

def show_warning_message(message):
    st.markdown(f'<div class="warning-message">⚠️ {message}</div>', unsafe_allow_html=True)

# --- Enhanced Progress Bar ---
def create_animated_progress_bar():
    progress_bar = st.progress(0)
    status_text = st.empty()
    return progress_bar, status_text

# --- Utility Functions (keeping original functionality) ---
def get_column(df, colname):
    """FIXED: Handle integer column names properly"""
    for col in df.columns:
        col_str = str(col).strip().lower()
        colname_str = str(colname).strip().lower()
        if col_str == colname_str:
            return col
    raise KeyError(f"Column '{colname}' not found. Available columns: {df.columns.tolist()}")

def get_raw_unique_names(series):
    return pd.Series(series).dropna().drop_duplicates().tolist()

def fix_tally_columns(df_tally):
    """Fix Tally sheet column structure when headers are wrong"""
    expected_cols = ['GSTIN of supplier', 'Supplier', 'Invoice number', 'Invoice Date',
                    'Invoice Value', 'Rate', 'Taxable Value', 'Integrated Tax',
                    'Central Tax', 'State/UT tax', 'Cess']
    
    if (len(df_tally.columns) >= 2 and
        str(df_tally.columns[0]).startswith('Unnamed') and
        not any(col.lower().strip() == 'supplier' for col in df_tally.columns)):
        
        new_columns = []
        for i, col in enumerate(df_tally.columns):
            if i < len(expected_cols):
                new_columns.append(expected_cols[i])
            else:
                new_columns.append(f"Column_{i}")
        df_tally.columns = new_columns
    
    return df_tally

def create_default_format():
    """Create default Excel format for users"""
    # Sample data for Tally sheet
    tally_data = {
        'GSTIN of supplier': ['27AABCU9603R1ZX', '27AABCU9603R1ZY', ''],
        'Supplier': ['ABC Private Ltd', 'XYZ Industries', 'GHI Enterprises'], 
        'Invoice number': ['INV-0001', 'INV-0002', 'INV-0003'],
        'Invoice Date': ['15-04-2024', '20-04-2024', '25-04-2024'],
        'Invoice Value': [118000, 177000, 112000],
        'Rate': [18, 18, 12],
        'Taxable Value': [100000, 150000, 100000],
        'Integrated Tax': [0, 27000, 0],
        'Central Tax': [9000, 0, 6000],
        'State/UT tax': [9000, 0, 6000],
        'Cess': [0, 0, 0]
    }
    
    # Sample data for GSTR-2A sheet
    gstr_data = {
        'GSTIN of supplier': ['27AABCU9603R1ZX', '27AABCU9603R1ZZ', ''],
        'Supplier': ['ABC Private Limited', 'DEF Corporation', 'GHI Enterprises'],
        'Invoice number': ['INV-0001', 'INV-0004', 'INV-0003'],
        'Invoice Date': ['15-04-2024', '25-04-2024', '25-04-2024'],
        'Invoice Value': [118000, 89600, 112000],
        'Rate': [18, 12, 12],
        'Taxable Value': [100000, 80000, 100000],
        'Integrated Tax': [0, 0, 0],
        'Central Tax': [9000, 4800, 6000],
        'State/UT tax': [9000, 4800, 6000],
        'Cess': [0, 0, 0]
    }
    
    df_tally = pd.DataFrame(tally_data)
    df_gstr = pd.DataFrame(gstr_data)
    
    # Create Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Add empty row for header space
        empty_row = pd.DataFrame([[''] * len(df_tally.columns)], columns=df_tally.columns)
        empty_row.to_excel(writer, sheet_name='Tally', index=False, header=False)
        df_tally.to_excel(writer, sheet_name='Tally', index=False, startrow=1)
        
        empty_row_gstr = pd.DataFrame([[''] * len(df_gstr.columns)], columns=df_gstr.columns)
        empty_row_gstr.to_excel(writer, sheet_name='GSTR-2A', index=False, header=False)
        df_gstr.to_excel(writer, sheet_name='GSTR-2A', index=False, startrow=1)
    
    output.seek(0)
    return output.getvalue()

def show_help_instructions():
    """Enhanced help instructions"""
    st.markdown("""
    <div class="results-container">
    
    ## 🔄 GST Reconciliation Help Guide

    ### 📋 Step-by-Step Process:

    **1. 📂 File Preparation:**
    - 📥 Download the default format template  
    - 📝 Fill your data in 'Tally' and 'GSTR-2A' sheets
    - 📤 Upload the completed Excel file

    **2. 🚀 Fuzzy Matching:**
    - 🎯 Set match threshold (80% recommended)
    - ▶️ Click "Start Matching" to find similar supplier names
    - ✅ Review matches and confirm with Yes/No dropdowns
    - 📊 Use bulk operations for quick confirmations
    - 📥 Download/upload for offline bulk editing
    - ▶️ Click "Continue" to save confirmations

    **3. 🔁 Name Replacement:**
    - 🔄 Replaces Tally supplier names with GSTR names
    - ✅ Only processes "Yes" confirmed matches
    - 📋 Creates 'Tally_Replaced' sheet

    **4. 📊 GST Reconciliation:**
    - 🧮 Compares GST amounts between books and GSTR-2A
    - 📈 Shows variances and summaries
    - 📊 Creates detailed comparison sheets

    **5. 🧾 Invoice Reconciliation:**
    - 📋 Invoice-wise detailed comparison
    - 🔍 Identifies missing invoices
    - 💰 Shows amount differences

    **6. 📥 Final Download:**
    - 📦 Download complete Excel with all results
    - 📂 Contains all analysis sheets in one file

    ### 💡 Tips:
    - 🎯 Higher threshold = stricter matching
    - ✅ Review fuzzy matches carefully
    - 📊 Check variance reports for discrepancies
    - 🔍 Use invoice reconciliation for detailed analysis

    ### ⚠️ Common Issues:
    - 📋 Ensure correct sheet names: 'Tally' and 'GSTR-2A'
    - 📝 Keep column headers as per template
    - 📅 Date format: DD-MM-YYYY or DD/MM/YYYY
    - 🔢 Numeric columns should contain only numbers
    
    </div>
    """, unsafe_allow_html=True)

# --- Enhanced Fuzzy Matching Logic ---
def two_way_match(tally_list, gstr_list, threshold):
    match_map, used_tally, used_gstr = {}, set(), set()
    tally_upper = {name.upper(): name for name in tally_list}
    gstr_upper = {name.upper(): name for name in gstr_list}
    tally_keys, gstr_keys = list(tally_upper.keys()), list(gstr_upper.keys())
    
    total_steps = len(gstr_keys) + len(tally_keys)
    progress_bar, status_text = create_animated_progress_bar()
    
    # GSTR to Tally matching
    for i, gstr_name in enumerate(gstr_keys):
        best_match, score = process.extractOne(gstr_name, tally_keys, scorer=fuzz.ratio)
        gstr_real = gstr_upper[gstr_name]
        
        if best_match and score >= threshold and best_match not in used_tally:
            tally_real = tally_upper[best_match]
            match_map[(gstr_real, tally_real)] = (gstr_real, tally_real, score)
            used_gstr.add(gstr_name)
            used_tally.add(best_match)
        else:
            match_map[(gstr_real, '')] = (gstr_real, '', 0)
            used_gstr.add(gstr_name)
        
        progress = (i + 1) / total_steps
        progress_bar.progress(progress)
        status_text.markdown(f'<div class="info-message">🔍 GSTR → Tally: {i+1}/{len(gstr_keys)}</div>', unsafe_allow_html=True)
        time.sleep(0.01)
    
    # Tally to GSTR matching
    for i, tally_name in enumerate(tally_keys):
        if tally_name in used_tally:
            continue
            
        best_match, score = process.extractOne(tally_name, gstr_keys, scorer=fuzz.ratio)
        tally_real = tally_upper[tally_name]
        
        if best_match and score >= threshold and best_match not in used_gstr:
            gstr_real = gstr_upper[best_match]
            match_map[(gstr_real, tally_real)] = (gstr_real, tally_real, score)
            used_tally.add(tally_name)
            used_gstr.add(best_match)
        else:
            match_map[('', tally_real)] = ('', tally_real, 0)
            used_tally.add(tally_name)
        
        progress = (len(gstr_keys) + i + 1) / total_steps
        progress_bar.progress(progress)
        status_text.markdown(f'<div class="info-message">🔍 Tally → GSTR: {i+1}/{len(tally_keys)}</div>', unsafe_allow_html=True)
        time.sleep(0.01)
    
    # Complete progress
    progress_bar.progress(1.0)
    status_text.markdown('<div class="success-message">🎉 Matching completed successfully!</div>', unsafe_allow_html=True)
    time.sleep(1)
    progress_bar.empty()
    status_text.empty()
    
    results = []
    for gstr_name, tally_name, score in match_map.values():
        # Set default confirmation based on match quality
        if gstr_name and tally_name and score >= 80:
            confirm = "Yes"
        else:
            confirm = "No"
        results.append([gstr_name, tally_name, score, confirm])
    
    return results

# --- Enhanced Display Functions ---
def display_dataframe_with_title(df, title, description=""):
    """Display dataframe with modern styling"""
    st.markdown(f"""
    <div class="results-container">
        <h3 class="column-header">{title}</h3>
        {f'<p style="color: #64748b; margin-bottom: 1rem;">{description}</p>' if description else ''}
    </div>
    """, unsafe_allow_html=True)
    st.dataframe(df, use_container_width=True)

# --- Enhanced Streamlit App ---
def show_reconciliation_tool():
    # Apply custom CSS first
    apply_custom_css()
    
    st.set_page_config(
        page_title="GSTR vs Tally Reconciliation",
        page_icon="📊",
        layout="wide"
    )
    
    st.markdown("""
    <div class="main-header">
        <h1>🔍 GST Reconciliation Tool</h1>
        <p>Advanced AI-powered reconciliation between Tally and GSTR-2A data</p>
    </div>
    """, unsafe_allow_html=True)
    # Initialize session state
    if 'uploaded_file' not in st.session_state:
        st.session_state.uploaded_file = None
    if 'match_results' not in st.session_state:
        st.session_state.match_results = None
    if 'temp_file_path' not in st.session_state:
        st.session_state.temp_file_path = None
    if 'matching_completed' not in st.session_state:
        st.session_state.matching_completed = False
    if 'manual_confirmations' not in st.session_state:
        st.session_state.manual_confirmations = {}
    if 'all_processes_completed' not in st.session_state:
        st.session_state.all_processes_completed = False
    if 'saved_match_results' not in st.session_state:
        st.session_state.saved_match_results = None
    if 'name_replacement_done' not in st.session_state:
        st.session_state.name_replacement_done = False
    if 'gst_reconciliation_done' not in st.session_state:
        st.session_state.gst_reconciliation_done = False
    if 'invoice_reconciliation_done' not in st.session_state:
        st.session_state.invoice_reconciliation_done = False

    # Top section with file upload and help
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        st.markdown('<h3 class="column-header">📂 Upload Excel File</h3>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader(
            "Choose an Excel file with 'Tally' and 'GSTR-2A' sheets",
            type=['xlsx', 'xls']
        )
    
    with col2:
        st.markdown('<h3 class="column-header">📥 Default Format</h3>', unsafe_allow_html=True)
        st.download_button(
            label="📥 Download Template",
            data=create_default_format(),
            file_name=f"GST_Reconciliation_Template_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download this template to see the required format",
            use_container_width=True
        )
        show_info_message("📋 Download this template to see the required Excel format with sample data")
    
    with col3:
        st.markdown('<h3 class="column-header">❓ Help</h3>', unsafe_allow_html=True)
        if st.button("❓ Show Instructions", use_container_width=True):
            st.session_state.show_help = not st.session_state.get('show_help', False)
    
    if st.session_state.get('show_help', False):
        with st.expander("📖 Help Guide", expanded=True):
            show_help_instructions()

    if uploaded_file is not None:
        st.session_state.uploaded_file = uploaded_file
        
        # Save uploaded file to temporary location
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            st.session_state.temp_file_path = tmp_file.name
        
        show_success_message(f"File uploaded: {uploaded_file.name}")
        
        # Validate sheets
        try:
            excel_file = pd.ExcelFile(st.session_state.temp_file_path)
            sheets = excel_file.sheet_names
            
            if 'Tally' in sheets and 'GSTR-2A' in sheets:
                show_success_message("Required sheets 'Tally' and 'GSTR-2A' found!")
            else:
                show_error_message(f"Required sheets not found. Available sheets: {sheets}")
                return
        except Exception as e:
            show_error_message(f"Error reading file: {e}")
            return

    if st.session_state.uploaded_file is None:
        show_info_message("👆 Please upload an Excel file to continue")
        return

    # Main functionality tabs with animations
    tab1, tab2, tab3, tab4 = st.tabs([
        "🚀 Fuzzy Matching",
        "🔁 Name Replacement", 
        "📊 GST Reconciliation",
        "🧾 Invoice Reconciliation"
    ])

    with tab1:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        st.header("🚀 Start Fuzzy Matching")
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            threshold = st.number_input(
                "Match Threshold (%)",
                min_value=0,
                max_value=100,
                value=80,
                help="Minimum similarity score for matching names"
            )
        
        if st.button("🚀 Start Matching", use_container_width=True):
            try:
                with st.spinner("🔄 Processing fuzzy matching..."):
                    # Read data
                    df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally', header=1)
                    df_gstr = pd.read_excel(st.session_state.temp_file_path, sheet_name='GSTR-2A', header=1)
                    
                    # Get supplier columns
                    col_supplier_tally = get_column(df_tally, 'Supplier')
                    col_supplier_gstr = get_column(df_gstr, 'Supplier')
                    
                    # Get unique names
                    tally_parties = get_raw_unique_names(df_tally[col_supplier_tally])
                    gstr_parties = get_raw_unique_names(df_gstr[col_supplier_gstr])
                    
                    # Perform matching with animation
                    matches = two_way_match(tally_parties, gstr_parties, threshold)
                    
                    st.session_state.match_results = matches
                    st.session_state.matching_completed = True
                    
                    # Initialize manual confirmations
                    st.session_state.manual_confirmations = {}
                    for i, match in enumerate(matches):
                        st.session_state.manual_confirmations[i] = match[3]  # Default confirmation
                
                show_success_message("Matching completed successfully!")
                
            except Exception as e:
                show_error_message(f"Error during matching: {e}")

        # Enhanced results display with animations
        if st.session_state.matching_completed and st.session_state.match_results:
            st.subheader("📋 Matching Results - Manual Confirmation")
            
            # Bulk operations with animated buttons
            st.write("**Bulk Operations:**")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("✅ Set All High Scores to Yes", help="Auto-confirm all matches with score ≥ 80%"):
                    for i, (gstr_name, tally_name, score, _) in enumerate(st.session_state.match_results):
                        if score >= 80 and gstr_name and tally_name:
                            st.session_state.manual_confirmations[i] = "Yes"
                        else:
                            st.session_state.manual_confirmations[i] = "No"
                    show_success_message("All high-score matches set to Yes!")
                    st.rerun()
            
            with col2:
                if st.button("❌ Set All to No"):
                    for i in range(len(st.session_state.match_results)):
                        st.session_state.manual_confirmations[i] = "No"
                    show_info_message("All matches set to No")
                    st.rerun()
            
            with col3:
                if st.button("🔄 Reset to Default"):
                    for i, match in enumerate(st.session_state.match_results):
                        st.session_state.manual_confirmations[i] = match[3]
                    show_info_message("Reset to default confirmations")
                    st.rerun()

            st.write("**Review and confirm matches below:**")
            
            # Create interactive confirmation interface with enhanced animations
            with st.form("confirmation_form"):
                cols = st.columns([3, 3, 1, 1.5])
                cols[0].write("**GSTR-2A Party**")
                cols[1].write("**Tally Party**")
                cols[2].write("**Score**")
                cols[3].write("**Confirmation**")
                
                for i, (gstr_name, tally_name, score, default_confirm) in enumerate(st.session_state.match_results):
                    cols = st.columns([3, 3, 1, 1.5])
                    
                    # Display names and score
                    cols[0].write(str(gstr_name) if gstr_name else "")
                    cols[1].write(str(tally_name) if tally_name else "")
                    cols[2].write(f"{score:.0f}%")
                    
                    # Dropdown for confirmation
                    current_value = st.session_state.manual_confirmations.get(i, default_confirm)
                    confirmation = cols[3].selectbox(
                        f"Confirm {i}",
                        options=["Yes", "No"],
                        index=0 if current_value == "Yes" else 1,
                        key=f"confirm_{i}",
                        label_visibility="collapsed"
                    )
                    st.session_state.manual_confirmations[i] = confirmation
                
                # Enhanced form submit button
                submitted = st.form_submit_button("🔄 Update All Confirmations", use_container_width=True)
                if submitted:
                    show_success_message("Confirmations updated!")

            # Excel Download/Upload functionality with enhanced styling
            st.subheader("📥📤 Excel Download/Upload for Bulk Editing")
            col1, col2 = st.columns(2)
            
            with col1:
                # Prepare current data for download
                current_results = []
                for i, (gstr_name, tally_name, score, _) in enumerate(st.session_state.match_results):
                    current_confirmation = st.session_state.manual_confirmations.get(i, "No")
                    current_results.append([gstr_name, tally_name, score, current_confirmation])
                
                df_download = pd.DataFrame(current_results,
                                         columns=['GSTR-2A Party', 'Tally Party', 'Score', 'Manual Confirmation'])
                
                # Create Excel file in memory
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_download.to_excel(writer, sheet_name='Matching_Results', index=False)
                output.seek(0)
                
                st.download_button(
                    label="📥 Download for Offline Editing",
                    data=output.getvalue(),
                    file_name=f"matching_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            with col2:
                # Upload edited file
                uploaded_confirmations = st.file_uploader(
                    "📤 Upload Edited Confirmations",
                    type=['xlsx', 'xls'],
                    help="Upload the Excel file with your edited confirmations"
                )
                
                if uploaded_confirmations is not None:
                    try:
                        df_uploaded = pd.read_excel(uploaded_confirmations)
                        # Validate columns
                        required_cols = ['GSTR-2A Party', 'Tally Party', 'Score', 'Manual Confirmation']
                        if all(col in df_uploaded.columns for col in required_cols):
                            # Update confirmations
                            for i, row in df_uploaded.iterrows():
                                if i < len(st.session_state.match_results):
                                    confirmation = str(row['Manual Confirmation']).strip()
                                    if confirmation.upper() in ['YES', 'Y', '1', 'TRUE']:
                                        st.session_state.manual_confirmations[i] = "Yes"
                                    else:
                                        st.session_state.manual_confirmations[i] = "No"
                            show_success_message("Confirmations updated from uploaded file!")
                            st.rerun()
                        else:
                            show_error_message("Invalid file format. Required columns: " + ", ".join(required_cols))
                    except Exception as e:
                        show_error_message(f"Error reading uploaded file: {e}")

            # Enhanced continue button
            st.subheader("▶️ Continue to Next Step")
            if st.button("▶️ Continue with Selected Confirmations", type="primary", use_container_width=True):
                try:
                    # Create final results with progress animation
                    with st.spinner("💾 Saving confirmations..."):
                        final_results = []
                        for i, (gstr_name, tally_name, score, _) in enumerate(st.session_state.match_results):
                            final_confirmation = st.session_state.manual_confirmations.get(i, "No")
                            final_results.append([gstr_name, tally_name, score, final_confirmation])
                        
                        # Create results dataframe
                        df_result = pd.DataFrame(final_results,
                                               columns=['GSTR-2A Party', 'Tally Party', 'Score', 'Manual Confirmation'])
                        df_result.sort_values(by=['Manual Confirmation', 'GSTR-2A Party', 'Tally Party'],
                                            ascending=[False, False, False], inplace=True)
                        
                        # Save to Excel with enhanced error handling
                        try:
                            book = load_workbook(st.session_state.temp_file_path)
                            if 'GSTR_Tally_Match' in book.sheetnames:
                                book.remove(book['GSTR_Tally_Match'])
                            
                            ws = book.create_sheet('GSTR_Tally_Match')
                            for r in dataframe_to_rows(df_result, index=False, header=True):
                                ws.append(r)
                            
                            for cell in ws[1]:
                                cell.font = Font(bold=True)
                            
                            book.save(st.session_state.temp_file_path)
                            book.close()
                            
                            show_success_message("Final confirmations saved successfully!")
                            
                        except Exception as e1:
                            try:
                                with pd.ExcelWriter(st.session_state.temp_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                                    df_result.to_excel(writer, sheet_name='GSTR_Tally_Match', index=False)
                                show_success_message("Final confirmations saved successfully!")
                            except Exception as e2:
                                show_error_message(f"Save failed: {e1}, {e2}")
                                return
                        
                        # Verify and show summary with celebration animation
                        st.session_state.saved_match_results = df_result
                        yes_count = sum(1 for conf in st.session_state.manual_confirmations.values() if conf == "Yes")
                        total_count = len(st.session_state.manual_confirmations)
                        
                        st.markdown('<div class="celebration">', unsafe_allow_html=True)
                        show_info_message(f"📊 Summary: {yes_count} out of {total_count} matches confirmed for replacement")
                        st.markdown('</div>', unsafe_allow_html=True)
                        
                        # Display final results with animation
                        display_dataframe_with_title(df_result, "📋 Final Matching Results", 
                                                   "These matches will be used for name replacement")
                        
                except Exception as e:
                    show_error_message(f"Error saving confirmations: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

    with tab2:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        st.header("🔁 Name Replacement")
        show_info_message("💡 First complete fuzzy matching, then use this to replace Tally names with GSTR names")
        
        if st.button("🔁 Replace Matched Names", use_container_width=True):
            try:
                with st.spinner("🔄 Processing name replacements..."):
                    # Progress simulation for visual feedback
                    progress_bar = st.progress(0)
                    for i in range(100):
                        time.sleep(0.01)
                        progress_bar.progress(i + 1)
                    progress_bar.empty()
                    
                    # Try to read match results from file first
                    df_matches = None
                    try:
                        df_matches = pd.read_excel(st.session_state.temp_file_path, sheet_name='GSTR_Tally_Match')
                        show_success_message("Match results loaded from Excel sheet")
                    except Exception as e1:
                        if 'saved_match_results' in st.session_state and st.session_state.saved_match_results is not None:
                            df_matches = st.session_state.saved_match_results
                            show_info_message("Match results loaded from session backup")
                        else:
                            show_error_message("Please complete fuzzy matching first and click 'Continue with Selected Confirmations'")
                            return

                    if df_matches is None or len(df_matches) == 0:
                        show_error_message("No match results found. Please complete fuzzy matching first.")
                        return

                    # Read Tally data
                    df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally', header=1)
                    col_supplier = get_column(df_tally, 'Supplier')

                    # Create name mapping
                    name_map = {}
                    replacement_count = 0
                    for _, row in df_matches.iterrows():
                        gstr_name = row['GSTR-2A Party']
                        tally_name = row['Tally Party']
                        confirmation = str(row['Manual Confirmation']).strip().upper()
                        
                        if confirmation == "YES" and pd.notna(gstr_name) and pd.notna(tally_name):
                            if gstr_name != '' and tally_name != '':
                                name_map[tally_name] = gstr_name
                                replacement_count += 1

                    # Apply replacements
                    df_new = df_tally.copy()
                    def replace_name(name):
                        if pd.isna(name):
                            return name
                        return name_map.get(name, name)

                    df_new[col_supplier] = df_new[col_supplier].apply(replace_name)

                    # Format Invoice Date
                    if 'Invoice Date' in df_new.columns:
                        df_new['Invoice Date'] = pd.to_datetime(df_new['Invoice Date'], errors='coerce').dt.strftime('%d-%m-%Y')

                    # Save updated data
                    with pd.ExcelWriter(st.session_state.temp_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_new.to_excel(writer, sheet_name='Tally_Replaced', index=False, header=True)

                    st.session_state.name_replacement_done = True
                    show_success_message(f"Replaced {replacement_count} supplier names successfully!")

                    # Show preview with animation
                    display_dataframe_with_title(df_new.head(10), "📋 Preview of Replaced Data", 
                                               "Showing first 10 rows of the replaced data")

            except Exception as e:
                show_error_message(f"Error during replacement: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

    with tab3:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        st.header("📊 GST Reconciliation")
        
        if st.button("📊 Run GST Reconciliation", use_container_width=True):
            try:
                with st.spinner("📊 Processing GST reconciliation..."):
                    # Enhanced progress bar
                    progress_bar = st.progress(0)
                    status = st.empty()
                    
                    # Check which Tally sheet to use
                    try:
                        df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally_Replaced', header=0)
                        tally_sheet_used = "Tally_Replaced"
                        show_info_message("Using Tally_Replaced sheet for reconciliation")
                    except:
                        df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally', header=1)
                        tally_sheet_used = "Tally"
                        show_info_message("Using original Tally sheet for reconciliation")
                    
                    progress_bar.progress(20)
                    status.markdown('<div class="info-message">📖 Loading GSTR-2A data...</div>', unsafe_allow_html=True)
                    
                    df_gstr = pd.read_excel(st.session_state.temp_file_path, sheet_name='GSTR-2A', header=1)
                    
                    progress_bar.progress(40)
                    status.markdown('<div class="info-message">🔧 Processing data...</div>', unsafe_allow_html=True)
                    
                    # Fix columns
                    df_tally = fix_tally_columns(df_tally)
                    
                    # Add Cess column if missing
                    for df in [df_tally, df_gstr]:
                        if 'Cess' not in df.columns:
                            df['Cess'] = 0
                    
                    progress_bar.progress(60)
                    status.markdown('<div class="info-message">🧮 Calculating reconciliation...</div>', unsafe_allow_html=True)
                    
                    # Get column mappings
                    col_name = get_column(df_tally, 'Supplier')
                    col_itax = get_column(df_tally, 'Integrated Tax')
                    col_ctax = get_column(df_tally, 'Central Tax')
                    col_stax = get_column(df_tally, 'State/UT tax')

                    # Check for GSTIN columns
                    try:
                        col_gstin_tally = get_column(df_tally, 'GSTIN of supplier')
                        col_gstin_gstr = get_column(df_gstr, 'GSTIN of supplier')
                        has_gstin = True
                    except KeyError:
                        has_gstin = False

                    # Process data
                    df_tally_processed = df_tally.copy()
                    df_gstr_processed = df_gstr.copy()
                    
                    progress_bar.progress(80)
                    status.markdown('<div class="info-message">📊 Creating comparisons...</div>', unsafe_allow_html=True)

                    if has_gstin:
                        # GSTIN-based grouping
                        df_tally_processed[col_gstin_tally] = df_tally_processed[col_gstin_tally].fillna('NO_GSTIN').astype(str)
                        df_gstr_processed[col_gstin_gstr] = df_gstr_processed[col_gstin_gstr].fillna('NO_GSTIN').astype(str)

                        def create_group_key(row, gstin_col, supplier_col):
                            gstin_val = str(row[gstin_col]).strip() if pd.notna(row[gstin_col]) else ''
                            supplier_val = str(row[supplier_col]).strip() if pd.notna(row[supplier_col]) else ''
                            if not gstin_val or gstin_val == 'NO_GSTIN' or gstin_val == 'nan':
                                return f"SUPPLIER_{supplier_val}"
                            else:
                                return f"GSTIN_{gstin_val}"

                        df_tally_processed['Group_Key'] = df_tally_processed.apply(
                            lambda row: create_group_key(row, col_gstin_tally, col_name), axis=1
                        )
                        df_gstr_processed['Group_Key'] = df_gstr_processed.apply(
                            lambda row: create_group_key(row, col_gstin_gstr, col_name), axis=1
                        )
                    else:
                        # Supplier-based grouping only
                        df_tally_processed['Group_Key'] = df_tally_processed[col_name].fillna('UNKNOWN')
                        df_gstr_processed['Group_Key'] = df_gstr_processed[col_name].fillna('UNKNOWN')

                    # Group and aggregate
                    group_cols = ['Group_Key']
                    df_tally_grp = df_tally_processed.groupby(group_cols).agg({
                        col_name: 'first',
                        col_itax: 'sum',
                        col_ctax: 'sum',
                        col_stax: 'sum',
                        'Cess': 'sum'
                    }).reset_index()

                    df_gstr_grp = df_gstr_processed.groupby(group_cols).agg({
                        col_name: 'first',
                        col_itax: 'sum',
                        col_ctax: 'sum',
                        col_stax: 'sum',
                        'Cess': 'sum'
                    }).reset_index()

                    # Create comparisons
                    df_combined = pd.merge(df_gstr_grp, df_tally_grp, on=group_cols, how='inner', suffixes=('_GSTR', '_Tally'))
                    df_combined['Integrated Tax Variance'] = df_combined[col_itax + '_GSTR'] - df_combined[col_itax + '_Tally']
                    df_combined['Central Tax Variance'] = df_combined[col_ctax + '_GSTR'] - df_combined[col_ctax + '_Tally']
                    df_combined['State/UT Tax Variance'] = df_combined[col_stax + '_GSTR'] - df_combined[col_stax + '_Tally']
                    df_combined['Cess Variance'] = df_combined['Cess_GSTR'] - df_combined['Cess_Tally']

                    # Find missing entries
                    df_combined_outer = pd.merge(df_gstr_grp, df_tally_grp, on=group_cols, how='outer', suffixes=('_GSTR', '_Tally'))
                    not_in_tally = df_combined_outer[df_combined_outer[col_itax + '_Tally'].isna()]
                    not_in_gstr = df_combined_outer[df_combined_outer[col_itax + '_GSTR'].isna()]

                    # Create summary
                    df_summary = pd.DataFrame({
                        'Particulars': [
                            'GST Input as per GSTR-2A Sheet',
                            f'GST Input as per {tally_sheet_used}',
                            'Variance (1-2)'
                        ],
                        'Integrated Tax': [
                            df_gstr_grp[col_itax].sum(),
                            df_tally_grp[col_itax].sum(),
                            df_gstr_grp[col_itax].sum() - df_tally_grp[col_itax].sum()
                        ],
                        'Central Tax': [
                            df_gstr_grp[col_ctax].sum(),
                            df_tally_grp[col_ctax].sum(),
                            df_gstr_grp[col_ctax].sum() - df_tally_grp[col_ctax].sum()
                        ],
                        'State/UT Tax': [
                            df_gstr_grp[col_stax].sum(),
                            df_tally_grp[col_stax].sum(),
                            df_gstr_grp[col_stax].sum() - df_tally_grp[col_stax].sum()
                        ],
                        'Cess': [
                            df_gstr_grp['Cess'].sum(),
                            df_tally_grp['Cess'].sum(),
                            df_gstr_grp['Cess'].sum() - df_tally_grp['Cess'].sum()
                        ]
                    })

                    progress_bar.progress(100)
                    status.markdown('<div class="success-message">💾 Saving results...</div>', unsafe_allow_html=True)

                    # Save results
                    with pd.ExcelWriter(st.session_state.temp_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_summary.to_excel(writer, sheet_name='GST_Input_Summary', index=False)
                        df_combined.to_excel(writer, sheet_name='T_vs_G-2A', index=False)
                        not_in_tally.to_excel(writer, sheet_name='N_I_T_B_I_G', index=False)
                        not_in_gstr.to_excel(writer, sheet_name='N_I_G_B_I_T', index=False)

                    progress_bar.empty()
                    status.empty()
                    
                    st.session_state.gst_reconciliation_done = True
                    show_success_message(f"GST Reconciliation completed using {tally_sheet_used}!")

                    # Display all results with modern styling
                    display_dataframe_with_title(df_summary, "📊 GST Input Summary", 
                                               "Overall comparison between GSTR-2A and Tally data")

                    display_dataframe_with_title(df_combined, "📋 Detailed Comparison (T_vs_G-2A)", 
                                               "Party-wise detailed variance analysis")

                    display_dataframe_with_title(not_in_tally, "🚫 Not in Tally but in GSTR (N_I_T_B_I_G)", 
                                               "Parties present in GSTR-2A but missing in Tally")

                    display_dataframe_with_title(not_in_gstr, "🚫 Not in GSTR but in Tally (N_I_G_B_I_T)", 
                                               "Parties present in Tally but missing in GSTR-2A")

            except Exception as e:
                show_error_message(f"Error during reconciliation: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

    with tab4:
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        st.header("🧾 Invoice-wise Reconciliation")
        
        if st.button("🧾 Run Invoice Reconciliation", use_container_width=True):
            try:
                with st.spinner("🧾 Processing invoice-wise reconciliation..."):
                    # Enhanced progress feedback
                    progress_bar = st.progress(0)
                    status = st.empty()
                    
                    # Check which Tally sheet to use
                    try:
                        df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally_Replaced', header=0)
                        tally_sheet_used = "Tally_Replaced"
                        df_tally = fix_tally_columns(df_tally)
                    except:
                        df_tally = pd.read_excel(st.session_state.temp_file_path, sheet_name='Tally', header=1)
                        tally_sheet_used = "Tally"
                    
                    progress_bar.progress(30)
                    status.markdown('<div class="info-message">📖 Loading GSTR-2A data...</div>', unsafe_allow_html=True)
                    
                    df_gstr = pd.read_excel(st.session_state.temp_file_path, sheet_name='GSTR-2A', header=1)

                    progress_bar.progress(50)
                    status.markdown('<div class="info-message">🔧 Cleaning data...</div>', unsafe_allow_html=True)
                    
                    # Clean columns
                    for df in [df_tally, df_gstr]:
                        df.columns = df.columns.str.strip()
                        if 'Cess' not in df.columns:
                            df['Cess'] = 0
                        df['GSTIN of supplier'] = df['GSTIN of supplier'].fillna('No GSTIN')

                    progress_bar.progress(70)
                    status.markdown('<div class="info-message">📊 Grouping invoices...</div>', unsafe_allow_html=True)
                    
                    # Group by invoice
                    group_columns = ['GSTIN of supplier', 'Supplier', 'Invoice number']
                    def consolidate(df):
                        return df.groupby(group_columns).agg({
                            'Taxable Value': 'sum',
                            'Integrated Tax': 'sum',
                            'Central Tax': 'sum',
                            'State/UT tax': 'sum',
                            'Cess': 'sum'
                        }).reset_index()

                    tally_grouped = consolidate(df_tally)
                    gstr_grouped = consolidate(df_gstr)

                    progress_bar.progress(90)
                    status.markdown('<div class="info-message">🧮 Calculating variances...</div>', unsafe_allow_html=True)
                    
                    # Combine and calculate variances
                    df_combined = pd.merge(gstr_grouped, tally_grouped, on=group_columns, how='outer', suffixes=('_GSTR', '_Tally')).fillna(0)
                    df_combined['Taxable Value Variance'] = df_combined['Taxable Value_GSTR'] - df_combined['Taxable Value_Tally']
                    df_combined['Integrated Tax Variance'] = df_combined['Integrated Tax_GSTR'] - df_combined['Integrated Tax_Tally']
                    df_combined['Central Tax Variance'] = df_combined['Central Tax_GSTR'] - df_combined['Central Tax_Tally']
                    df_combined['State/UT Tax Variance'] = df_combined['State/UT tax_GSTR'] - df_combined['State/UT tax_Tally']
                    df_combined['Cess Variance'] = df_combined['Cess_GSTR'] - df_combined['Cess_Tally']

                    progress_bar.progress(100)
                    status.markdown('<div class="success-message">💾 Saving results...</div>', unsafe_allow_html=True)
                    
                    # Save to Excel
                    with pd.ExcelWriter(st.session_state.temp_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                        df_combined.to_excel(writer, sheet_name='Invoice_Recon', index=False)

                    progress_bar.empty()
                    status.empty()
                    
                    st.session_state.invoice_reconciliation_done = True
                    show_success_message(f"Invoice-wise reconciliation completed using {tally_sheet_used}!")

                    # Mark all processes as completed
                    st.session_state.all_processes_completed = True

                    # Display results with animation
                    display_dataframe_with_title(df_combined, "📋 Invoice-wise Reconciliation Results", 
                                               "Detailed invoice-wise comparison with variances")

            except Exception as e:
                show_error_message(f"Error during invoice reconciliation: {e}")
        
        st.markdown('</div>', unsafe_allow_html=True)

    # Enhanced Final Download Section with Process Status
    if (st.session_state.matching_completed or
        st.session_state.get('temp_file_path') and
        os.path.exists(st.session_state.temp_file_path)):
        
        st.markdown("---")
        st.markdown('<div class="results-container">', unsafe_allow_html=True)
        st.header("📦 Final Downloads & Process Status")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown('<h3 class="column-header">📥 Complete Reconciliation Report</h3>', unsafe_allow_html=True)
            show_info_message("📋 Download the complete Excel file with all sheets and analysis results")
            
            try:
                with open(st.session_state.temp_file_path, 'rb') as file:
                    st.download_button(
                        label="📥 Download Complete Excel Report",
                        data=file.read(),
                        file_name=f"Complete_GST_Reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary"
                    )
            except:
                show_warning_message("⚠️ Complete file not available. Please run at least one process first.")
        
        with col2:
            st.markdown('<h3 class="column-header">📊 Process Status</h3>', unsafe_allow_html=True)
            
            try:
                excel_file = pd.ExcelFile(st.session_state.temp_file_path)
                available_sheets = excel_file.sheet_names
                
                # Enhanced process status with persistent tracking
                process_status = [
                    {
                        "name": "🚀 Fuzzy Matching",
                        "sheet": "GSTR_Tally_Match",
                        "session_var": "matching_completed",
                        "description": "Supplier name matching completed"
                    },
                    {
                        "name": "🔁 Name Replacement", 
                        "sheet": "Tally_Replaced",
                        "session_var": "name_replacement_done",
                        "description": "Tally names replaced with GSTR names"
                    },
                    {
                        "name": "📊 GST Reconciliation",
                        "sheet": "GST_Input_Summary",
                        "session_var": "gst_reconciliation_done", 
                        "description": "GST amounts reconciled and analyzed"
                    },
                    {
                        "name": "🧾 Invoice Reconciliation",
                        "sheet": "Invoice_Recon",
                        "session_var": "invoice_reconciliation_done",
                        "description": "Invoice-wise comparison completed"
                    }
                ]
                
                completed_count = 0
                
                for process in process_status:
                    sheet_exists = process["sheet"] in available_sheets
                    session_completed = st.session_state.get(process["session_var"], False)
                    
                    if sheet_exists or session_completed:
                        st.markdown(f'''
                        <div class="status-card">
                            <strong>{process["name"]}: ✅ Completed</strong><br>
                            <small style="color: #64748b;">{process["description"]}</small>
                        </div>
                        ''', unsafe_allow_html=True)
                        completed_count += 1
                        
                        # Update session state based on sheet existence
                        if sheet_exists:
                            st.session_state[process["session_var"]] = True
                    else:
                        st.markdown(f'''
                        <div style="background: rgba(239, 68, 68, 0.1); border-left: 4px solid #ef4444; padding: 1rem; margin: 0.5rem 0; border-radius: 8px;">
                            <strong>{process["name"]}: ❌ Not Done</strong><br>
                            <small style="color: #64748b;">{process["description"]}</small>
                        </div>
                        ''', unsafe_allow_html=True)
                
                # Show completion percentage
                completion_percentage = (completed_count / len(process_status)) * 100
                st.markdown(f'''
                <div style="background: linear-gradient(135deg, #06b6d4 0%, #3b82f6 100%); color: white; padding: 1rem; border-radius: 12px; text-align: center; margin: 1rem 0;">
                    <strong>Overall Progress: {completion_percentage:.0f}% Complete</strong><br>
                    <small>{completed_count} out of {len(process_status)} processes completed</small>
                </div>
                ''', unsafe_allow_html=True)
                
                # Show available sheets
                st.markdown('<h4 style="margin-top: 2rem;">📂 Available Sheets in Excel:</h4>', unsafe_allow_html=True)
                for i, sheet in enumerate(available_sheets, 1):
                    st.markdown(f'''
                    <div style="background: rgba(255, 255, 255, 0.9); padding: 0.5rem 1rem; margin: 0.25rem 0; border-radius: 8px; border-left: 3px solid #10b981;">
                        <strong>{i}. {sheet}</strong>
                    </div>
                    ''', unsafe_allow_html=True)
                    
            except Exception as e:
                show_error_message(f"Error checking file status: {e}")

        # Celebration animation if all processes completed
        if completed_count == 4:  # All 4 processes completed
            st.session_state.all_processes_completed = True
            st.markdown('''
            <div class="celebration" style="text-align: center; margin: 2rem 0;">
                <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; padding: 2rem; border-radius: 20px; box-shadow: 0 10px 30px rgba(16, 185, 129, 0.3);">
                    <h2 style="margin: 0; font-size: 2rem;">🎉 Congratulations! 🎉</h2>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1.2rem;">All reconciliation processes completed successfully!</p>
                    <p style="margin: 0.5rem 0 0 0;">You can now download the complete report with all analysis results.</p>
                </div>
            </div>
            ''', unsafe_allow_html=True)
        
        st.markdown('</div>', unsafe_allow_html=True)

def main_with_navigation():
    """Enhanced main function with navigation"""
    
    # Page configuration
    st.set_page_config(
        page_title="GST Reconciliation Tool",
        page_icon="🔍",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Sidebar navigation
    st.sidebar.title("🔍 GST Reconciliation Tool")
    st.sidebar.markdown("---")
    
    # Navigation menu
    page = st.sidebar.selectbox("📋 Navigate to:", [
        "🏠 Home - Reconciliation Tool",
        "📄 About Us", 
        "🔒 Privacy Policy",
        "✉️ Contact Us"
    ])
    
    # Page routing
    if page == "🏠 Home - Reconciliation Tool":
        show_reconciliation_tool()
    elif page == "📄 About Us":
        show_about_page()
    elif page == "🔒 Privacy Policy":
        show_privacy_policy()
    elif page == "✉️ Contact Us":
        show_contact_page()

def create_required_files():
    """Create required directories and files"""
    pages_dir = os.path.join(current_dir, 'pages')
    utils_dir = os.path.join(current_dir, 'utils')
    
    os.makedirs(pages_dir, exist_ok=True)
    os.makedirs(utils_dir, exist_ok=True)
    
    for directory in [pages_dir, utils_dir]:
        init_file = os.path.join(directory, '__init__.py')
        if not os.path.exists(init_file):
            with open(init_file, 'w') as f:
                f.write("# This file makes the directory a Python package\n")

if __name__ == "__main__":
    create_required_files()
    main_with_navigation()
